# sets openpyxl alias to oxl
import openpyxl as oxl
# imports Path class from pathlib module to access the workbook files
from pathlib import Path
# Necessary functions for the cretiong of the bar chart
from openpyxl.chart import BarChart, Reference

def update_workbook(filename):
    # saves given filename passed in as workbook
    workbook = oxl.load_workbook(filename)
    # stores the workbook sheet in 'sheet'
    sheet = workbook['Sheet1']
    # Loop runs through each of the sales figures, generates the commission, and stores in a new cell
    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        commission = cell.value * 0.1

        new_cell = sheet.cell(row, 4)
        new_cell.value = commission
    # creates boundaries for the bar chart accessible values
    values = Reference(sheet,
                       min_row=3,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    # forms the bar chart and saves the changes to the workbook
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'F3')
    workbook.save(filename)
# accesses the targeted path(s) to retrieve the specific workbooks to update
# targets the sales directory. Can be left blank to target the current directory.
path = Path("sales")
# *.xlsx searches for all xl files in 'sales' directory
for file in path.glob('*.xlsx'):
    update_workbook(file)
